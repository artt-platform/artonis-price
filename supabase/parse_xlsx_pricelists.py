"""Parse 2 known XLSX price lists → insert into Supabase price_observations."""
import os, sys, re, requests, unicodedata
import openpyxl
sys.path.insert(0,'/Users/trietnguyen/Documents/Company/Artonis/App/ArtonisArtistPrice')
os.chdir('/Users/trietnguyen/Documents/Company/Artonis/App/ArtonisArtistPrice')
from artonis_price_mvp import normalize_key

ENV={}
for line in open('.env.local'):
    line=line.strip()
    if line and not line.startswith('#') and '=' in line:
        k,v=line.split('=',1); ENV[k]=v
URL=ENV['SUPABASE_URL']; KEY=ENV['SUPABASE_SERVICE_ROLE_KEY']
H={'apikey':KEY,'Authorization':f'Bearer {KEY}','Content-Type':'application/json','Prefer':'return=minimal'}

DRY = os.environ.get('DRY','0')=='1'

exhs = requests.get(f'{URL}/rest/v1/exhibitions?select=id,title,artists_text,drive_path', headers={'apikey':KEY}).json()
arts = requests.get(f'{URL}/rest/v1/artists?select=id,name,normalized_name', headers={'apikey':KEY}).json()
artist_by_norm = {a['normalized_name']: a['id'] for a in arts}

def find_exh_artist(folder_part):
    """Match a folder substring to exhibition. Return (exh_id, artist_id)."""
    for e in exhs:
        if folder_part.lower() in (e.get('drive_path') or '').lower():
            art_text = (e.get('artists_text') or '').strip()
            names = [n.strip() for n in re.split(r'[,\n+]', art_text) if n.strip()]
            aid = artist_by_norm.get(normalize_key(names[0])) if len(names)==1 else None
            return e['id'], aid, e.get('title')
    return None, None, None

def parse_dimensions(s):
    if not s: return ''
    s = str(s).replace('×','x').replace('X','x')
    return s.strip()

# === Bùi Tiến Tuấn — DANH SÁCH & BÁO GIÁ.xlsx ===
def parse_bui_tien_tuan():
    path = '/tmp/artonis_drive/Bùi Tiến Tuấn - Lụa Là - 251030 - Thái Công Art Gallery/DANH SÁCH & BÁO GIÁ.xlsx'
    wb = openpyxl.load_workbook(path, data_only=True)
    exh_id, aid, title = find_exh_artist('Bùi Tiến Tuấn - Lụa Là')
    print(f'Bùi Tiến Tuấn Lụa Là: exh={exh_id}, artist={aid}, title={title!r}')
    rows_out = []
    # Use Sheet1 (has more columns, status notes); fall back to Sheet2 if needed
    ws = wb['Sheet1']
    for row in ws.iter_rows(min_row=4, values_only=True):
        # cols: stt | _ | tên | chất liệu | kích thước | năm | giá USD | NOTE | ...
        if not row[0] or not row[2]:
            continue
        title_aw, medium, dims, year, price, note = row[2], row[3], row[4], row[5], row[6], row[7]
        if not isinstance(price, (int,float)) or price <= 0:
            continue
        status = 'sold' if (note and 'sold' in str(note).lower()) else ''
        rows_out.append({
            'artist_id': aid, 'exhibition_id': exh_id,
            'artwork_title': str(title_aw).strip()[:300],
            'medium': str(medium).strip() if medium else None,
            'dimensions': parse_dimensions(dims),
            'year': str(int(year)) if isinstance(year,(int,float)) else (str(year) if year else None),
            'price_amount': float(price), 'currency': 'USD',
            'status': status, 'confidence': 0.9,
        })
    return rows_out

# === Noah Bùi — Một Mảnh Thiên Đường ===
def parse_noah_bui():
    path = '/tmp/artonis_drive/Noah Bùi - Một Mảnh Thiên Đường - 250616 - Chillala/Noah Bùi - Một Mảnh Thiên Đường - 20250616 - Chillala (1).xlsx'
    wb = openpyxl.load_workbook(path, data_only=True)
    exh_id, aid, title = find_exh_artist('Noah Bùi - Một Mảnh Thiên Đường')
    print(f'Noah Bùi: exh={exh_id}, artist={aid}, title={title!r}')
    rows_out = []
    ws = wb['Sheet1']
    for row in ws.iter_rows(min_row=4, values_only=True):
        # cols: No | Title | Photo | Dimensions | Medium | Date | Price USD | _
        if not row[0] or not row[1]:
            continue
        title_aw, dims, medium, year, price = row[1], row[3], row[4], row[5], row[6]
        if not isinstance(price, (int,float)) or price <= 0:
            continue
        rows_out.append({
            'artist_id': aid, 'exhibition_id': exh_id,
            'artwork_title': str(title_aw).strip()[:300],
            'medium': str(medium).strip() if medium else None,
            'dimensions': parse_dimensions(dims),
            'year': str(int(year)) if isinstance(year,(int,float)) else (str(year) if year else None),
            'price_amount': float(price), 'currency': 'USD',
            'status': '', 'confidence': 0.9,
        })
    return rows_out

all_rows = parse_bui_tien_tuan() + parse_noah_bui()
print(f'\nTotal rows: {len(all_rows)}')
for r in all_rows[:5]:
    print(f'  artist={r["artist_id"]} exh={r["exhibition_id"]} | {r["artwork_title"][:35]} | {r["price_amount"]} {r["currency"]}')
if not DRY and all_rows:
    rsp = requests.post(f'{URL}/rest/v1/price_observations', headers=H, json=all_rows, timeout=60)
    print(f'\ninsert HTTP {rsp.status_code}')
    if rsp.status_code not in (200,201,204):
        print(rsp.text[:400])
